#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
修复版的Markdown文件插入脚本，改进中文内容的处理
将结果保存在output\toolify_analysis_当日日期\文件夹下，保持与之前相同的格式和样式
"""

import os
import sys
import re
import glob
import pandas as pd
import traceback
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def markdown_to_plaintext(markdown_content, language="cn"):
    """
    将Markdown格式的文本转换为适合Excel单元格的纯文本格式，
    支持中文和英文内容的正确处理

    Args:
        markdown_content (str): Markdown格式的文本内容
        language (str): 语言，"cn"或"en"

    Returns:
        str: 转换后的纯文本内容
    """
    if not markdown_content:
        return ""

    # 替换所有表格相关符号
    text = re.sub(r'\|', ' | ', markdown_content)

    # 替换标题格式（# 标题）
    text = re.sub(r'^\s*#{1,6}\s+(.+)$', r'\1', text, flags=re.MULTILINE)

    # 替换加粗和斜体格式
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)  # 加粗
    text = re.sub(r'__(.+?)__', r'\1', text)      # 加粗
    text = re.sub(r'\*(.+?)\*', r'\1', text)      # 斜体
    text = re.sub(r'_(.+?)_', r'\1', text)        # 斜体

    # 替换列表格式（- 和 * 开头的项）
    text = re.sub(r'^\s*[-*]\s+(.+)$', r'• \1', text, flags=re.MULTILINE)

    # 替换链接格式 [文本](链接)
    text = re.sub(r'\[(.+?)\]\((.+?)\)', r'\1 (\2)', text)

    # 替换图片格式 ![alt](链接)
    text = re.sub(r'!\[(.+?)\]\((.+?)\)', r'[图片: \1]', text)

    # 移除水平线
    text = re.sub(r'^\s*[-*_]{3,}\s*$', '', text, flags=re.MULTILINE)

    # 移除表格格式符号（| --- | 类的行）
    text = re.sub(r'^\s*\|[-:\s|]+\|\s*$', '', text, flags=re.MULTILINE)

    # 替换多个连续空行为单个空行
    text = re.sub(r'\n{3,}', '\n\n', text)

    # 根据语言替换英文标题为中文标题
    if language == "cn":
        # 产品信息部分
        text = re.sub(r'Product Information', r'产品信息', text)
        text = re.sub(r'## 产品信息', r'产品信息', text)

        # 产品分析框架部分
        text = re.sub(r'Product Analysis Framework', r'产品分析框架', text)
        text = re.sub(r'## 产品分析框架', r'产品分析框架', text)

        # SWOT分析部分
        text = re.sub(r'SWOT Analysis', r'SWOT分析', text)
        text = re.sub(r'## SWOT分析', r'SWOT分析', text)

        # 评分体系部分
        text = re.sub(r'Rating System', r'评分体系', text)
        text = re.sub(r'## 评分体系', r'评分体系', text)

        # 关键洞察与建议部分
        text = re.sub(r'Key Insights and Recommendations', r'关键洞察与建议', text)
        text = re.sub(r'## 关键洞察与建议', r'关键洞察与建议', text)

        # 替换英文标签为中文标签
        text = re.sub(r'📊 Rank:', r'📊 排名:', text)
        text = re.sub(r'💰 Revenue:', r'💰 收入:', text)
        text = re.sub(r'🔗 Product Link:', r'🔗 产品链接:', text)
        text = re.sub(r'🔍 Analysis Link:', r'🔍 分析链接:', text)
        text = re.sub(r'👀 Monthly Visits:', r'👀 月访问量:', text)
        text = re.sub(r'🏢 Company:', r'🏢 公司:', text)
        text = re.sub(r'🗓️ Founded Year:', r'🗓️ 成立日期:', text)
        text = re.sub(r'💲 Pricing:', r'💲 定价:', text)
        text = re.sub(r'📱 Platform:', r'📱 平台:', text)
        text = re.sub(r'🔧 Core Features:', r'🔧 核心功能:', text)
        text = re.sub(r'🌐 Use Cases:', r'🌐 应用场景:', text)
        text = re.sub(r'⏱️ Analysis Time:', r'⏱️ 分析时间:', text)
        text = re.sub(r'🤖 Analysis Tool:', r'🤖 分析工具:', text)

    return text

def format_excel_with_beauty(df, output_file):
    """用美观的格式保存Excel文件"""
    try:
        # 创建Excel写入器
        writer = pd.ExcelWriter(output_file, engine='openpyxl')

        # 写入数据
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 应用格式（使用openpyxl的样式）
        # 设置列宽
        for idx, col in enumerate(df.columns):
            column_letter = get_column_letter(idx + 1)
            # 标准列宽
            col_width = 15

            # 特定列宽度调整
            if col in ["产品分析", "完整分析"]:
                col_width = 50  # 分析列宽一些
            elif col in ["收入", "Revenue", "年收入", "Annual Revenue"]:
                col_width = 12

            worksheet.column_dimensions[column_letter].width = col_width

        # 设置标题行格式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 应用标题行格式
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置单元格边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用到所有单元格
        for row in range(1, len(df) + 2):  # +2 因为有标题行和索引从1开始
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col_num)
                cell.border = thin_border

                # 内容单元格格式
                if row > 1:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 特殊列格式
                    column_name = df.columns[col_num-1]
                    if column_name in ["产品分析", "完整分析"]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 保存文件
        writer.close()
        print(f"✅ Excel文件已保存: {output_file}")
        return True
    except Exception as e:
        print(f"❌ 保存Excel格式化文件失败: {str(e)}")
        return False

def insert_md_to_excel(excel_file=None, markdown_dir=None, date_str=None, language="cn"):
    """
    将Markdown文件内容插入到Excel表格中

    Args:
        excel_file (str, optional): Excel文件路径，如果不提供则自动查找
        markdown_dir (str, optional): Markdown文件目录，如果不提供则自动查找
        date_str (str, optional): 日期字符串，如果不提供则使用当前日期
        language (str): 语言，"cn"或"en"

    Returns:
        bool: 是否成功
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    # 如果没有提供Excel文件路径，尝试查找
    if excel_file is None:
        data_dir = os.path.join("output", "toolify_data")
        excel_file = os.path.join(data_dir, f"Toolify_Top_AI_Revenue_Rankings_{language.upper()}_{date_str}.xlsx")

        if not os.path.exists(excel_file):
            # 尝试查找其他可能的文件名
            alt_excel_file = os.path.join(data_dir, f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx")
            if os.path.exists(alt_excel_file):
                excel_file = alt_excel_file
            else:
                # 尝试查找任何匹配的Excel文件
                excel_files = glob.glob(os.path.join(data_dir, f"*{language.upper()}*{date_str}*.xlsx"))
                if excel_files:
                    excel_file = excel_files[0]
                else:
                    print(f"错误: 找不到Excel文件，请手动指定 --excel-file 参数")
                    return False

    # 如果没有提供Markdown目录，尝试查找
    if markdown_dir is None:
        markdown_dir = os.path.join("output", f"toolify_analysis_{date_str}", language, "markdown_files")
        if not os.path.exists(markdown_dir):
            print(f"错误: 找不到Markdown文件目录: {markdown_dir}")
            print("请手动指定 --markdown-dir 参数")
            return False

    # 检查文件和目录是否存在
    if not os.path.exists(excel_file):
        print(f"错误: Excel文件不存在: {excel_file}")
        return False

    if not os.path.exists(markdown_dir):
        print(f"错误: Markdown目录不存在: {markdown_dir}")
        return False

    # 读取Excel文件
    try:
        print(f"正在读取Excel文件: {excel_file}")
        df = pd.read_excel(excel_file)
        print(f"已加载 {len(df)} 条产品数据")
        print(f"Excel列名: {df.columns.tolist()}")

        # 确定排名列名
        rank_column = None
        for possible_name in ["Ranking", "排名", "Rank"]:
            if possible_name in df.columns:
                rank_column = possible_name
                print(f"找到排名列: {rank_column}")
                break

        if not rank_column:
            print("错误: 无法在Excel中找到排名列")
            return False

        # 将DataFrame转换为字典列表
        tools = df.to_dict('records')
    except Exception as e:
        print(f"错误: 无法读取Excel文件: {str(e)}")
        traceback.print_exc()
        return False

    # 获取所有Markdown文件
    md_files = glob.glob(os.path.join(markdown_dir, "*.md"))
    if not md_files:
        print(f"错误: 在 {markdown_dir} 中找不到任何Markdown文件")
        return False

    print(f"找到 {len(md_files)} 个Markdown文件")

    # 准备分析结果
    analysis_results = []
    for md_file in md_files:
        # 从文件名提取排名
        file_name = os.path.basename(md_file)
        rank_match = re.match(r'^(\d+)-', file_name)
        if not rank_match:
            print(f"警告: 无法从文件名提取排名: {file_name}")
            continue

        rank = int(rank_match.group(1))
        print(f"处理排名 {rank} 的产品: {file_name}")

        # 从原始数据中找到对应的产品
        product = None
        for p in tools:
            try:
                p_rank = p.get(rank_column)
                if isinstance(p_rank, str):
                    p_rank = int(p_rank)

                if p_rank == rank:
                    product = p
                    break
            except (ValueError, TypeError):
                continue

        if not product:
            print(f"警告: 找不到排名为 {rank} 的产品数据")
            continue

        # 从文件内容中检测分析工具
        api_name = "DeepSeek AI"  # 默认值
        try:
            with open(md_file, 'r', encoding='utf-8') as f:
                content = f.read()
                # 检测是否包含OpenAI或GPT
                if "OpenAI" in content or "GPT" in content:
                    api_name = "OpenAI GPT-4"
                # 检测是否包含DeepSeek
                elif "DeepSeek" in content:
                    api_name = "DeepSeek AI"
                print(f"  检测到分析工具: {api_name}")
        except Exception as e:
            print(f"  警告: 无法读取文件内容: {str(e)}")

        # 添加到分析结果列表
        analysis_results.append({
            "product": product,
            "markdown_path": md_file,
            "api_name": api_name  # 添加分析工具信息
        })

    if not analysis_results:
        print("错误: 没有找到有效的分析结果")
        return False

    print(f"准备更新 {len(analysis_results)} 个产品的分析结果")

    # 更新Excel文件
    try:
        # 确保移除产品分析列
        if "产品分析" in df.columns:
            df = df.drop(columns=["产品分析"])

        # 只保留完整分析列
        if "完整分析" not in df.columns:
            df["完整分析"] = ""

        # 更新分析结果
        total_results = len(analysis_results)
        print(f"\n开始更新Excel文件 [共{total_results}个产品]")

        # 定义进度条颜色和符号
        bar_length = 40
        filled_char = "█"  # 实心方块
        empty_char = "░"   # 空心方块

        for i, result in enumerate(analysis_results):
            # 计算进度
            progress = (i + 1) / total_results
            filled_length = int(bar_length * progress)

            # 清除当前行并显示新进度条
            if i > 0:  # 不是第一个元素
                print("\r", end="")

            # 打印进度条
            progress_bar = filled_char * filled_length + empty_char * (bar_length - filled_length)
            print(f"\r进度: {progress_bar} {progress:.1%} ({i+1}/{total_results})", end="")

            product = result["product"]

            # 查找对应的行
            mask = df[rank_column] == product[rank_column]

            if not mask.any():
                print(f"\n警告: 无法在Excel中找到排名为 {product[rank_column]} 的产品")
                continue

            # 获取分析工具信息
            api_name = result.get("api_name", "DeepSeek AI")  # 默认使用DeepSeek AI

            # 更新完整分析列
            if result["markdown_path"] and os.path.exists(result["markdown_path"]):
                try:
                    with open(result["markdown_path"], 'r', encoding='utf-8') as md_file:
                        markdown_content = md_file.read()

                        # 更新分析工具信息
                        if "🤖 分析工具: " in markdown_content:
                            markdown_content = re.sub(
                                r'🤖 分析工具: [^\n]+',
                                f'🤖 分析工具: {api_name}',
                                markdown_content
                            )
                        elif "🤖 Analysis Tool: " in markdown_content:
                            markdown_content = re.sub(
                                r'🤖 Analysis Tool: [^\n]+',
                                f'🤖 Analysis Tool: {api_name}',
                                markdown_content
                            )

                        # 转换为纯文本并更新Excel
                        plain_text = markdown_to_plaintext(markdown_content, language)
                        df.loc[mask, "完整分析"] = plain_text

                        # 将更新后的内容写回文件
                        with open(result["markdown_path"], 'w', encoding='utf-8') as md_out:
                            md_out.write(markdown_content)

                        print(f"\r已更新排名 {product[rank_column]} 的产品分析", end="")
                except Exception as md_error:
                    print(f"\n⚠️ 读取markdown文件失败: {str(md_error)}")

        # 完成后打印完成信息
        print(f"\r进度: {filled_char * bar_length} 100.0% ({total_results}/{total_results}) ✅ 完成!\n")

        # 直接使用output/toolify_data目录作为输出目录
        output_dir = os.path.join("output", "toolify_data")

        # 确保目录存在
        os.makedirs(output_dir, exist_ok=True)

        print(f"\n将结果保存到目录: {output_dir}")

        # 生成输出文件名
        file_name = os.path.basename(excel_file)
        base_name, ext = os.path.splitext(file_name)
        output_file = os.path.join(output_dir, f"{base_name}_analyzed{ext}")

        # 保存Excel文件
        format_excel_with_beauty(df, output_file)

        print(f"✅ 已将分析结果保存到Excel文件: {output_file}")
        return output_file

    except Exception as e:
        print(f"❌ 更新Excel文件失败: {str(e)}")
        traceback.print_exc()
        return None

def main():
    # 获取当前日期
    date_str = "20250422"  # 使用固定日期以匹配现有文件

    # 设置默认路径
    excel_file = os.path.join("output", "toolify_data", f"Toolify_AI_Revenue_CN_{date_str}.xlsx")
    markdown_dir = os.path.join("output", f"toolify_analysis_{date_str}", "cn", "markdown_files")

    print("\n=== 环境检查 ===")
    print(f"当前工作目录: {os.getcwd()}")
    print(f"Excel文件路径: {excel_file}")
    print(f"Markdown目录: {markdown_dir}")

    # 检查文件和目录是否存在
    if not os.path.exists(excel_file):
        # 尝试查找其他可能的文件名
        alt_excel_file = os.path.join("output", "toolify_data", f"Toolify_Top_AI_Revenue_Rankings_CN_{date_str}.xlsx")
        if os.path.exists(alt_excel_file):
            excel_file = alt_excel_file
            print(f"找到替代Excel文件: {excel_file}")
        else:
            print(f"错误: Excel文件不存在: {excel_file}")
            print(f"替代文件也不存在: {alt_excel_file}")
            return False

    if not os.path.exists(markdown_dir):
        print(f"错误: Markdown目录不存在: {markdown_dir}")
        # 列出父目录内容
        parent_dir = os.path.dirname(markdown_dir)
        if os.path.exists(parent_dir):
            print(f"\n父目录 {parent_dir} 中的内容:")
            for item in os.listdir(parent_dir):
                print(f"  - {item}")
        return False

    # 列出Markdown目录中的文件
    md_files = [f for f in os.listdir(markdown_dir) if f.endswith('.md')]
    print(f"\n找到 {len(md_files)} 个Markdown文件:")
    for md_file in md_files[:5]:  # 只显示前5个文件
        print(f"  - {md_file}")
    if len(md_files) > 5:
        print(f"  ... 还有 {len(md_files)-5} 个文件")

    print("\n=== 开始处理 ===")
    # 执行插入操作
    success = insert_md_to_excel(
        excel_file=excel_file,
        markdown_dir=markdown_dir,
        date_str=date_str,
        language="cn"
    )

    if success:
        print("\n✅ 成功完成MD文件插入到Excel")
    else:
        print("\n❌ MD文件插入失败")

    return success

if __name__ == "__main__":
    sys.exit(0 if main() else 1)
