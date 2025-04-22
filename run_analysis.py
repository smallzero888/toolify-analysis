#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
运行数据采集和分析的脚本

此脚本执行以下操作:
1. 运行toolify_scraper.py爬取数据（中文和英文）
2. 运行analyze_data_with_gpu进行数据分析
3. 运行product_analyzer.py进行产品分析（可选）
4. 支持分析指定排名范围的产品（集成自analyze_ranks.py）

使用:
    # 基本用法
    python run_analysis.py [--scraping] [--analysis] [--analyze-products]

    # 分析指定排名范围的产品
    python run_analysis.py --rank-range 1-5 --language cn  # 分析排名1-5的中文产品
    python run_analysis.py --rank-range 6-10 --language en  # 分析排名6-10的英文产品

注意:
    本脚本已集成analyze_ranks.py的功能，不再需要单独运行该脚本。
"""

import os
import sys
import argparse
import subprocess
import traceback
import time
import re
import glob
from datetime import datetime

# 检查是否有必要的模块
try:
    import pandas as pd
    import numpy as np
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"错误: 缺少必要的模块: {e}")
    print("请先运行: pip install pandas numpy openpyxl")
    sys.exit(1)

# 导入工具模块
try:
    from toolify_utils import (
        TF_AVAILABLE,
        GPU_AVAILABLE,
        find_latest_files
    )

    # 导入爬虫模块的函数
    from toolify_scraper import scrape_toolify_ranking, main as scrape_main
    print("✅ 已成功导入工具模块")
except ImportError as e:
    print(f"错误: 无法导入工具模块: {e}")
    traceback.print_exc()
    sys.exit(1)


def run_scraping(output_dir="output/toolify_data", languages=None):
    """
    运行网站爬虫获取数据

    Args:
        output_dir (str): 输出目录
        languages (list): 要爬取的语言列表，例如 ["cn", "en"]

    Returns:
        dict: 包含爬取结果的字典
    """
    if languages is None:
        languages = ["cn", "en"]

    results = {}
    date_str = datetime.now().strftime("%Y%m%d")

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    for lang in languages:
        try:
            print(f"\n🔍 开始爬取 {lang.upper()} 数据...")

            # 设置输出目录
            date_str = datetime.now().strftime("%Y%m%d")
            # 直接使用output_dir，不创建子目录
            os.makedirs(output_dir, exist_ok=True)

            # 为不同语言设置不同的URL
            if lang.lower() == "cn":
                url = "https://www.toolify.ai/zh/Best-AI-Tools-revenue"
                # 将cn转换为zh，以与toolify_scraper.py保持一致
                scraper_lang = "zh"
            else:  # en
                url = "https://www.toolify.ai/Best-AI-Tools-revenue"
                scraper_lang = "en"

            # 设置输出文件名
            output_file = os.path.join(output_dir, f"Toolify_AI_Revenue_{lang.upper()}_{date_str}.xlsx")

            # 爬取数据
            data = scrape_toolify_ranking(url, output_file, language=scraper_lang)

            if data:
                rows_count = len(data)
                print(f"✅ 已成功爬取 {rows_count} 条{lang.upper()}数据并保存到 {output_file}")
                results[lang] = {
                    "count": rows_count,
                    "file": output_file,
                    "data": data
                }
            else:
                print(f"⚠️ 爬取{lang.upper()}数据失败或没有数据")

        except Exception as e:
            print(f"❌ 爬取{lang.upper()}数据时出错: {str(e)}")
            traceback.print_exc()

    return results


# 定义基本的数据分析功能
def analyze_data_with_gpu(cn_data=None, en_data=None, cn_file=None, en_file=None, output_dir="output/toolify_analysis", date_str=None):
    """
    使用GPU加速分析数据（如果可用）

    Args:
        cn_data (list, optional): 中文数据列表
        en_data (list, optional): 英文数据列表
        cn_file (str, optional): 中文数据文件路径
        en_file (str, optional): 英文数据文件路径
        output_dir (str): 输出目录
        date_str (str, optional): 日期字符串

    Returns:
        dict: 分析结果
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    print(f"🔍 开始数据分析，使用GPU: {GPU_AVAILABLE}")

    # 如果提供了文件路径但没有数据，尝试加载数据
    if not cn_data and cn_file and os.path.exists(cn_file):
        try:
            cn_data = pd.read_excel(cn_file).to_dict('records')
            print(f"✅ 从文件加载中文数据: {cn_file}")
        except Exception as e:
            print(f"⚠️ 无法读取中文数据文件: {str(e)}")

    if not en_data and en_file and os.path.exists(en_file):
        try:
            en_data = pd.read_excel(en_file).to_dict('records')
            print(f"✅ 从文件加载英文数据: {en_file}")
        except Exception as e:
            print(f"⚠️ 无法读取英文数据文件: {str(e)}")

    # 确保至少有一个数据集
    if not cn_data and not en_data:
        print("⚠️ 没有数据可供分析")
        return None

    # 分析逻辑
    try:
        # 合并中英文数据
        all_data = []
        if cn_data:
            all_data.extend(cn_data)
        if en_data:
            all_data.extend(en_data)

        # 提取并清理月访问量数据
        monthly_visits = []
        for item in all_data:
            visit_str = item.get("Monthly Visits") or item.get("月访问量", "0")
            try:
                # 移除非数字字符
                visit_str = visit_str.replace(",", "").replace("K", "000").replace("M", "000000")
                visits = float(visit_str)
                monthly_visits.append(visits)
            except (ValueError, AttributeError):
                monthly_visits.append(0)

        # 使用NumPy进行基本统计分析
        stats = {
            "count": len(monthly_visits),
            "sum": np.sum(monthly_visits),
            "mean": np.mean(monthly_visits),
            "median": np.median(monthly_visits),
            "std": np.std(monthly_visits),
            "min": np.min(monthly_visits),
            "max": np.max(monthly_visits)
        }

        # 创建分析输出文件
        os.makedirs(output_dir, exist_ok=True)
        summary_file = os.path.join(output_dir, f"Traffic_Summary_{date_str}.xlsx")

        # 将分析结果保存为DataFrame
        stats_df = pd.DataFrame([stats])
        stats_df.to_excel(summary_file, index=False)

        print(f"✅ 分析完成，生成报告: {summary_file}")
        return {
            "stats": stats,
            "summary_file": summary_file
        }

    except Exception as e:
        print(f"❌ 数据分析出错: {str(e)}")
        traceback.print_exc()
        return None


def run_analysis(scraping_results=None, output_dir="output/toolify_analysis", date_str=None):
    """
    运行数据分析

    Args:
        scraping_results (dict): 爬虫结果
        output_dir (str): 输出目录
        date_str (str): 日期字符串

    Returns:
        dict: 分析结果
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    # 如果没有提供爬虫结果，尝试查找最新的文件
    if not scraping_results:
        print("\n🔍 寻找最新的数据文件...")

        try:
            latest_files = find_latest_files(output_dir, "Toolify_AI_Revenue_*.xlsx")

            if not latest_files:
                print(f"⚠️ 在{output_dir}目录中未找到数据文件")
                return None

            # 重建结果结构
            scraping_results = {}
            for lang, file_path in latest_files.items():
                if os.path.exists(file_path):
                    try:
                        data = pd.read_excel(file_path).to_dict('records')
                        scraping_results[lang] = {
                            "count": len(data),
                            "file": file_path,
                            "data": data
                        }
                        print(f"✅ 已加载{lang.upper()}数据: {file_path}")
                    except Exception as e:
                        print(f"⚠️ 无法读取{file_path}: {str(e)}")
                else:
                    print(f"⚠️ 文件不存在: {file_path}")

        except Exception as e:
            print(f"❌ 查找数据文件时出错: {str(e)}")
            traceback.print_exc()
            return None

    # 运行分析
    print("\n🔍 开始分析数据...")

    try:
        # 提取数据和文件路径
        cn_data = scraping_results.get("cn", {}).get("data")
        en_data = scraping_results.get("en", {}).get("data")
        cn_file = scraping_results.get("cn", {}).get("file")
        en_file = scraping_results.get("en", {}).get("file")

        if not cn_data and not en_data and not cn_file and not en_file:
            print("⚠️ 没有找到可分析的数据")
            return None

        # 运行GPU加速分析
        analysis_result = analyze_data_with_gpu(
            cn_data=cn_data,
            en_data=en_data,
            cn_file=cn_file,
            en_file=en_file,
            output_dir=output_dir,
            date_str=date_str
        )

        return analysis_result

    except Exception as e:
        print(f"❌ 数据分析时出错: {str(e)}")
        traceback.print_exc()
        return None


def run_product_analyzer(data_files, output_dir="output/toolify_analysis", batch_size=5, use_gpu=False, count=None):
    """
    运行产品分析

    Args:
        data_files (dict): 数据文件字典，格式为 {"cn": "path/to/file.xlsx", "en": "path/to/file.xlsx"}
        output_dir (str): 输出目录
        batch_size (int): 批处理大小
        use_gpu (bool): 是否使用GPU
        count (int, optional): 要分析的产品数量

    Returns:
        bool: 是否成功
    """
    # 验证输入
    if not data_files:
        print("⚠️ 未提供数据文件")
        return False

    # 查找product_analyzer.py
    script_path = "product_analyzer.py"
    if not os.path.exists(script_path):
        print(f"❌ 未找到产品分析脚本: {script_path}")
        return False

    success = True

    for lang, file_path in data_files.items():
        if not os.path.exists(file_path):
            print(f"⚠️ 文件不存在: {file_path}")
            continue

        try:
            print(f"\n🔍 开始分析 {lang.upper()} 产品数据: {file_path}")

            # 构建命令
            cmd = [
                sys.executable,
                script_path,
                "-i", file_path,
                "-o", f"{output_dir}/{lang}",
                "-b", str(batch_size),
                "-l", lang
            ]

            # 添加产品数量参数（如果提供）
            if count is not None:
                cmd.extend(["-c", str(count)])

            # 添加GPU标志（如果启用）
            if use_gpu and GPU_AVAILABLE:
                cmd.append("--gpu")

            # 运行命令
            print(f"🚀 执行命令: {' '.join(cmd)}")
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

            # 实时输出
            while True:
                # 添加安全检查防止访问None
                if process.stdout is None:
                    break

                output = process.stdout.readline()
                if output == '' and process.poll() is not None:
                    break
                if output:
                    print(output.strip())

            # 获取返回码
            return_code = process.poll()

            if return_code == 0:
                print(f"✅ {lang.upper()}产品分析完成")
            else:
                print(f"❌ {lang.upper()}产品分析失败，返回码: {return_code}")
                success = False

                # 打印错误输出
                if process.stderr is not None:
                    stderr = process.stderr.read()
                    if stderr:
                        print(f"错误信息:\n{stderr}")

        except Exception as e:
            print(f"❌ 运行产品分析时出错: {str(e)}")
            traceback.print_exc()
            success = False

    return success


# 新增: 分析指定产品ID的函数
def analyze_specific_products(product_ids, language="cn", date_str=None, api="deepseek", use_gpu=False, update_excel=False, retry_count=3, excel_file=None):
    """
    分析指定产品ID的产品

    Args:
        product_ids (str): 产品ID列表，如"1,2,3"
        language (str): 语言，"cn"或"en"
        date_str (str, optional): 日期字符串
        api (str): 使用的API，"deepseek"或"openai"
        use_gpu (bool): 是否使用GPU
        update_excel (bool): 是否将分析结果插入到Excel表格中
        retry_count (int): API调用失败时的重试次数
        excel_file (str, optional): Excel文件路径，如果不提供则自动查找

    Returns:
        bool: 是否成功
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    # 解析产品ID列表
    try:
        ids = [int(id.strip()) for id in product_ids.split(",") if id.strip()]
        if not ids:
            print("错误: 未提供有效的产品ID")
            return False
    except ValueError:
        print("错误: 产品ID必须是数字，以逗号分隔")
        return False

    # 确定输入文件
    # 如果指定了excel_file参数，使用指定的文件
    if 'excel_file' in globals() and excel_file and os.path.exists(excel_file):
        input_file = excel_file
    else:
        # 尝试在不同的目录中查找文件
        possible_paths = [
            # 标准路径
            os.path.join("output", "toolify_data", f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx"),
            # 备用路径
            os.path.join("output", f"toolify_data", f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx"),
            # 兼容旧路径
            os.path.join("output", "toolify_data", f"toolify_analysis_{date_str}", f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx")
        ]

        input_file = None
        for path in possible_paths:
            if os.path.exists(path):
                input_file = path
                break

        if not input_file:
            print(f"错误: 找不到输入文件")
            print("尝试了以下路径:")
            for path in possible_paths:
                print(f"  - {path}")
            print("请确保已经爬取了榜单数据或使用--excel-file参数指定文件路径")
            return False

    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        products = df.to_dict('records')
        print(f"已加载 {len(products)} 条产品数据")
    except Exception as e:
        print(f"错误: 无法读取Excel文件: {str(e)}")
        return False

    # 查找对应ID的产品
    products_to_analyze = []
    for id in ids:
        # 查找对应ID的产品（ID对应Excel中的索引+1）
        if 0 < id <= len(products):
            products_to_analyze.append(products[id-1])
        else:
            print(f"警告: 找不到ID为 {id} 的产品，有效ID范围: 1-{len(products)}")

    if not products_to_analyze:
        print("错误: 没有找到任何有效的产品ID")
        return False

    # 设置输出目录
    output_dir = os.path.join("output", f"toolify_analysis_{date_str}", language, "markdown_files")

    # 构建命令
    cmd = [
        "python", "analyze_product.py"
    ]

    # 添加API参数
    if api.lower() == "openai":
        cmd.extend(["--api", "openai"])

    # 添加语言参数
    cmd.extend(["--language", language])

    # 添加日期参数
    cmd.extend(["--date", date_str])

    # 显示即将执行的操作
    lang_display = "中文" if language == "cn" else "英文"
    print(f"正在分析{lang_display}榜单中ID为 {', '.join(map(str, ids))} 的产品...")

    # 逐个分析产品
    success = True
    for product in products_to_analyze:
        # 获取产品排名
        rank = product.get("Rank") or product.get("排名")
        if not rank:
            # 如果没有排名，使用索引位置+1作为排名
            rank = products.index(product) + 1

        # 构建分析单个产品的命令
        product_cmd = cmd.copy()
        product_cmd.extend(["--rank", str(rank)])

        # 执行命令
        attempt = 0
        product_success = False
        max_attempts = max(1, retry_count + 1)  # 至少尝试一次

        # 检查API密钥是否存在
        if api.lower() == "openai" and not os.environ.get("OPENAI_API_KEY"):
            print(f"\n错误: 未设置OPENAI_API_KEY环境变量，无法使用OpenAI API")
            print("请在.env文件中添加OPENAI_API_KEY=your_key或者切换到DeepSeek API")
            return False
        elif api.lower() == "deepseek" and not os.environ.get("DEEPSEEK_API_KEY"):
            print(f"\n错误: 未设置DEEPSEEK_API_KEY环境变量，无法使用DeepSeek API")
            print("请在.env文件中添加DEEPSEEK_API_KEY=your_key或者切换到OpenAI API")
            return False

        while attempt < max_attempts and not product_success:
            attempt += 1
            try:
                print(f"\n尝试 {attempt}/{max_attempts}: 正在分析产品 {rank}...")
                result = subprocess.run(product_cmd, capture_output=True, text=True)

                if result.returncode == 0:
                    print(f"分析产品 {rank} 完成!")
                    product_success = True
                else:
                    print(f"错误: 分析产品 {rank} 失败 (返回码: {result.returncode})")
                    print(f"错误输出: {result.stderr}")

                    if attempt < max_attempts:
                        wait_time = 5 * attempt  # 每次重试等待时间增加
                        print(f"将在 {wait_time} 秒后重试...")
                        time.sleep(wait_time)
            except Exception as e:
                print(f"错误: 执行分析时出错: {str(e)}")

                if attempt < max_attempts:
                    wait_time = 5 * attempt
                    print(f"将在 {wait_time} 秒后重试...")
                    time.sleep(wait_time)

        if not product_success:
            success = False

    # 如果成功并需要更新Excel，则将MD内容插入到Excel表格中
    if success and update_excel:
        try:
            # 获取所有分析结果
            markdown_dir = output_dir  # 输出目录已经包含了language
            if not os.path.exists(markdown_dir):
                print(f"警告: 找不到Markdown文件目录: {markdown_dir}")
                return success

            # 获取所有分析结果文件
            md_files = glob.glob(os.path.join(markdown_dir, "*.md"))
            if not md_files:
                print("警告: 找不到任何Markdown文件")
                return success

            print(f"\n找到 {len(md_files)} 个分析结果文件")

            # 读取原始数据
            try:
                # 读取原始数据文件
                tools = pd.read_excel(input_file).to_dict('records')
                print(f"\n已加载原始数据: {len(tools)} 条记录")
            except Exception as e:
                print(f"错误: 无法读取原始数据文件: {str(e)}")
                return success

            # 准备分析结果
            analysis_results = []
            for md_file in md_files:
                # 从文件名提取排名
                file_name = os.path.basename(md_file)
                rank_match = re.match(r'^(\d+)-', file_name)
                if not rank_match:
                    continue

                rank = int(rank_match.group(1))

                # 从原始数据中找到对应的产品
                product = None
                for p in tools:
                    p_rank = p.get("Rank") or p.get("排名")
                    try:
                        p_rank = int(p_rank)
                    except (ValueError, TypeError):
                        continue

                    if p_rank == rank:
                        product = p
                        break

                if not product:
                    continue

                # 添加到分析结果列表
                analysis_results.append({
                    "product": product,
                    "markdown_path": md_file
                })

            if not analysis_results:
                print("警告: 没有找到有效的分析结果")
                return success

            # 更新Excel文件
            from toolify_utils import update_excel_with_analysis
            updated_file = update_excel_with_analysis(
                input_file,  # 使用输入文件作为要更新的Excel文件
                analysis_results,
                markdown_dir=markdown_dir
            )

            if updated_file:
                print(f"\n成功更新Excel文件: {updated_file}")
            else:
                print("警告: 更新Excel文件失败")
        except Exception as e:
            print(f"错误: 更新Excel文件时出错: {str(e)}")
            traceback.print_exc()

    return success

# 分析指定排名范围的产品函数
def analyze_specific_ranks(rank_range, language="cn", date_str=None, api="deepseek", use_gpu=False, update_excel=False, retry_count=3):
    """
    分析指定排名范围的产品

    Args:
        rank_range (str): 排名范围，如"1-5"
        language (str): 语言，"cn"或"en"
        date_str (str, optional): 日期字符串
        api (str): 使用的API，"deepseek"或"openai"
        use_gpu (bool): 是否使用GPU
        update_excel (bool): 是否将分析结果插入到Excel表格中
        retry_count (int): API调用失败时的重试次数

    Returns:
        bool: 是否成功
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    # 检查排名范围格式
    if "-" not in rank_range:
        print("错误: 排名范围格式不正确，应为如'1-5'这样的格式")
        return False

    # 解析排名范围
    try:
        start_rank, end_rank = map(int, rank_range.split("-"))
    except ValueError:
        print("错误: 排名必须是数字")
        return False

    if start_rank <= 0 or end_rank <= 0:
        print("错误: 排名必须大于0")
        return False

    if start_rank > end_rank:
        print("错误: 起始排名不能大于结束排名")
        return False

    # 确定输入文件
    data_dir = os.path.join("output", "toolify_data")
    input_file = os.path.join(data_dir, f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx")

    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件 {input_file}")
        print("请确保已经爬取了榜单数据")
        return False

    # 计算参数
    start_index = start_rank - 1  # 转换为基于0的索引
    count = end_rank - start_rank + 1

    # 设置输出目录
    output_dir = os.path.join("output", f"toolify_analysis_{date_str}", language, "markdown_files")

    # 构建命令
    cmd = [
        "python", "product_analyzer.py",
        "-i", input_file,
        "-o", output_dir,
        "-s", str(start_index),
        "-c", str(count),
        "-l", language
    ]

    # 添加API参数
    if api.lower() == "openai":
        cmd.extend(["--api", "openai"])

    # 添加GPU标志（如果启用）
    if use_gpu and GPU_AVAILABLE:
        cmd.append("--gpu")

    # 显示即将执行的操作
    lang_display = "中文" if language == "cn" else "英文"
    print(f"正在分析{lang_display}榜单中排名{start_rank}到{end_rank}的产品...")

    # 执行命令
    success = False
    attempt = 0
    max_attempts = max(1, retry_count + 1)  # 至少尝试一次

    # 检查API密钥是否存在
    if api.lower() == "openai" and not os.environ.get("OPENAI_API_KEY"):
        print(f"\n错误: 未设置OPENAI_API_KEY环境变量，无法使用OpenAI API")
        print("请在.env文件中添加OPENAI_API_KEY=your_key或者切换到DeepSeek API")
        return False
    elif api.lower() == "deepseek" and not os.environ.get("DEEPSEEK_API_KEY"):
        print(f"\n错误: 未设置DEEPSEEK_API_KEY环境变量，无法使用DeepSeek API")
        print("请在.env文件中添加DEEPSEEK_API_KEY=your_key或者切换到OpenAI API")
        return False

    while attempt < max_attempts and not success:
        attempt += 1
        try:
            print(f"\n尝试 {attempt}/{max_attempts}: 正在分析产品...")
            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                print(f"\n分析完成! 结果已保存到 {output_dir} 目录下。")
                success = True
            else:
                print(f"错误: 分析运行失败 (返回码: {result.returncode})")
                print(f"错误输出: {result.stderr}")

                if attempt < max_attempts:
                    wait_time = 5 * attempt  # 每次重试等待时间增加
                    print(f"将在 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
        except Exception as e:
            print(f"错误: 执行分析时出错: {str(e)}")

            if attempt < max_attempts:
                wait_time = 5 * attempt
                print(f"将在 {wait_time} 秒后重试...")
                time.sleep(wait_time)

    # 如果成功并需要更新Excel，则将MD内容插入到Excel表格中
    if success and update_excel:
        try:
            # 获取所有分析结果
            markdown_dir = output_dir  # 输出目录已经包含了language
            if not os.path.exists(markdown_dir):
                print(f"警告: 找不到Markdown文件目录: {markdown_dir}")
                return success

            # 获取所有分析结果文件
            md_files = glob.glob(os.path.join(markdown_dir, "*.md"))
            if not md_files:
                print("警告: 找不到任何Markdown文件")
                return success

            print(f"\n找到 {len(md_files)} 个分析结果文件")

            # 读取原始数据
            try:
                # 读取原始数据文件
                tools = pd.read_excel(input_file).to_dict('records')
                print(f"\n已加载原始数据: {len(tools)} 条记录")
            except Exception as e:
                print(f"错误: 无法读取原始数据文件: {str(e)}")
                return success

            # 准备分析结果
            analysis_results = []
            for md_file in md_files:
                # 从文件名提取排名
                file_name = os.path.basename(md_file)
                rank_match = re.match(r'^(\d+)-', file_name)
                if not rank_match:
                    continue

                rank = int(rank_match.group(1))

                # 从原始数据中找到对应的产品
                product = None
                for p in tools:
                    p_rank = p.get("Rank") or p.get("排名")
                    try:
                        p_rank = int(p_rank)
                    except (ValueError, TypeError):
                        continue

                    if p_rank == rank:
                        product = p
                        break

                if not product:
                    continue

                # 添加到分析结果列表
                analysis_results.append({
                    "product": product,
                    "markdown_path": md_file
                })

            if not analysis_results:
                print("警告: 没有找到有效的分析结果")
                return success

            # 更新Excel文件
            from toolify_utils import update_excel_with_analysis
            updated_file = update_excel_with_analysis(
                input_file,  # 使用输入文件作为要更新的Excel文件
                analysis_results,
                markdown_dir=markdown_dir
            )

            if updated_file:
                print(f"\n成功更新Excel文件: {updated_file}")
            else:
                print("警告: 更新Excel文件失败")
        except Exception as e:
            print(f"错误: 更新Excel文件时出错: {str(e)}")
            traceback.print_exc()

    return success


def markdown_to_plaintext(md_text):
    """
    将Markdown文本转换为纯文本

    Args:
        md_text (str): Markdown文本

    Returns:
        str: 转换后的纯文本
    """
    # 移除代码块
    md_text = re.sub(r'```[\s\S]*?```', '', md_text)

    # 移除图片
    md_text = re.sub(r'!\[.*?\]\(.*?\)', '', md_text)

    # 移除链接，保留链接文本
    md_text = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', md_text)

    # 移除标题符号
    md_text = re.sub(r'^#{1,6}\s+', '', md_text, flags=re.MULTILINE)

    # 移除粗体和斜体
    md_text = re.sub(r'\*\*(.*?)\*\*', r'\1', md_text)
    md_text = re.sub(r'__(.*?)__', r'\1', md_text)
    md_text = re.sub(r'\*(.*?)\*', r'\1', md_text)
    md_text = re.sub(r'_(.*?)_', r'\1', md_text)

    # 移除引用符号
    md_text = re.sub(r'^>\s+', '', md_text, flags=re.MULTILINE)

    # 移除水平线
    md_text = re.sub(r'^-{3,}$', '', md_text, flags=re.MULTILINE)
    md_text = re.sub(r'^\*{3,}$', '', md_text, flags=re.MULTILINE)

    # 移除列表符号
    md_text = re.sub(r'^\s*[\*\-\+]\s+', '• ', md_text, flags=re.MULTILINE)
    md_text = re.sub(r'^\s*\d+\.\s+', '• ', md_text, flags=re.MULTILINE)

    # 移除多余的空行
    md_text = re.sub(r'\n{3,}', '\n\n', md_text)

    return md_text.strip()


def format_excel_with_beauty(wb, sheet, column_name="完整分析"):
    """
    美化Excel表格

    Args:
        wb: openpyxl工作簿对象
        sheet: openpyxl工作表对象
        column_name (str): 要格式化的列名

    Returns:
        None
    """
    # 定义样式
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列宽
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15

    # 查找分析结果列
    analysis_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == column_name:
            analysis_col = col
            break

    if analysis_col:
        # 设置分析结果列的宽度
        column_letter = get_column_letter(analysis_col)
        sheet.column_dimensions[column_letter].width = 80

    # 设置行高
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 300

    # 设置标题行样式
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置数据行样式
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

            # 为分析结果列设置特殊样式
            if col == analysis_col:
                cell.font = Font(name="微软雅黑", size=10)


def insert_md_to_excel(excel_file=None, markdown_dir=None, date_str=None, language="cn"):
    """
    将Markdown文件内容插入到Excel表格中

    Args:
        excel_file (str, optional): Excel文件路径，如果不提供则自动查找
        markdown_dir (str, optional): Markdown文件目录，如果不提供则自动查找
        date_str (str, optional): 日期字符串，如果不提供则使用当前日期
        language (str): 语言，"cn"或"en"

    Returns:
        bool: 是否成功
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    print(f"\n🔍 开始将Markdown文件插入到Excel表格中...")

    # 使用try-except块包装整个函数
    try:
        # 如果没有提供Excel文件路径，尝试查找最新的文件
        if not excel_file:
            data_dir = os.path.join("output", "toolify_data")
            excel_pattern = f"Toolify_AI_Revenue_{language.upper()}_{date_str}.xlsx"
            excel_files = glob.glob(os.path.join(data_dir, excel_pattern))

            if not excel_files:
                print(f"⚠️ 在{data_dir}目录中未找到匹配的Excel文件: {excel_pattern}")
                return False

            excel_file = excel_files[0]
            print(f"找到Excel文件: {excel_file}")

        # 如果未提供Markdown目录，尝试查找
        if not markdown_dir:
            md_base_dir = os.path.join("output", f"toolify_analysis_{date_str}")
            possible_md_dirs = [
                # 标准目录结构
                os.path.join(md_base_dir, language, "markdown_files"),
                # 兼容旧的分析目录
                os.path.join("output", "toolify_data", f"analysis_{date_str}", language, "markdown_files")
            ]

            print(f"正在查找Markdown目录...")
            for dir_path in possible_md_dirs:
                print(f"  检查目录: {dir_path}")
                if os.path.exists(dir_path) and os.path.isdir(dir_path):
                    # 检查目录中是否有.md文件
                    md_files = glob.glob(os.path.join(dir_path, "*.md"))
                    if md_files:
                        markdown_dir = dir_path
                        print(f"  找到Markdown目录: {markdown_dir} (包含 {len(md_files)} 个.md文件)")
                        break
                    else:
                        print(f"  目录存在但没有.md文件: {dir_path}")
                else:
                    print(f"  目录不存在: {dir_path}")

            if not markdown_dir:
                print(f"⚠️ 未找到Markdown目录，尝试了以下路径:")
                for dir_path in possible_md_dirs:
                    print(f"  - {dir_path}")
                return False

        # 读取Excel文件
        df = pd.read_excel(excel_file)

        # 确保完整分析列存在
        if "完整分析" not in df.columns:
            df["完整分析"] = ""
            print(f"已添加完整分析列")

        # 查找Markdown文件
        md_files = glob.glob(os.path.join(markdown_dir, "*.md"))
        if not md_files:
            print(f"⚠️ 在{markdown_dir}目录中未找到Markdown文件")
            return False

        # 为每个产品查找对应的Markdown文件
        for index, row in df.iterrows():
            # 获取产品ID（索引+1）
            product_id = index + 1

            # 查找对应的Markdown文件
            md_file = None
            for file in md_files:
                file_name = os.path.basename(file)
                # 尝试不同的模式匹配
                if f"_{product_id}_" in file_name:
                    md_file = file
                    break
                # 尝试匹配排名-开头的文件
                rank_match = re.match(r'^(\d+)-', file_name)
                if rank_match and int(rank_match.group(1)) == product_id:
                    md_file = file
                    break

            if md_file:
                # 读取Markdown文件内容
                try:
                    with open(md_file, "r", encoding="utf-8") as f:
                        md_content = f.read()
                except UnicodeDecodeError:
                    # 尝试使用其他编码
                    try:
                        with open(md_file, "r", encoding="gbk") as f:
                            md_content = f.read()
                    except UnicodeDecodeError:
                        print(f"\u2757 无法读取文件: {md_file}")
                        continue

                # 将Markdown内容转换为纯文本
                plain_text = markdown_to_plaintext(md_content)

                # 将Markdown内容添加到DataFrame中
                # 确保完整分析列存在
                if "完整分析" not in df.columns:
                    df["完整分析"] = ""

                # 处理可能的乱码问题
                try:
                    # 尝试直接使用纯文本
                    df.at[index, "完整分析"] = plain_text
                    print(f"✅ 已插入产品ID {product_id} 的分析结果")
                except Exception as e:
                    # 如果出错，尝试使用原始Markdown内容
                    print(f"❗ 插入纯文本时出错: {str(e)}")
                    try:
                        df.at[index, "完整分析"] = md_content
                        print(f"✅ 已插入产品ID {product_id} 的原始Markdown内容")
                    except Exception as e2:
                        print(f"❗ 插入原始Markdown内容时出错: {str(e2)}")

        # 保存更新后的Excel文件
        # 创建输出目录
        output_dir = os.path.join("output", f"toolify_analysis_{date_str}")
        os.makedirs(output_dir, exist_ok=True)

        # 生成输出文件名 - 使用固定命名格式而不是时间戳
        file_name = os.path.basename(excel_file)
        standard_output_file = os.path.join(output_dir, file_name.replace(".xlsx", "_analyzed.xlsx"))

        # 尝试保存文件
        try:
            df.to_excel(standard_output_file, index=False)
            output_file = standard_output_file
        except PermissionError:
            # 如果出现权限错误，尝试使用带时间戳的文件名
            timestamp = datetime.now().strftime("%H%M%S")
            backup_output_file = os.path.join(output_dir, file_name.replace(".xlsx", f"_analyzed_backup.xlsx"))
            print(f"标准文件被占用，尝试使用备用文件名: {backup_output_file}")
            try:
                df.to_excel(backup_output_file, index=False)
                output_file = backup_output_file
            except PermissionError:
                # 如果备用文件也被占用，使用带时间戳的文件名
                timestamp_output_file = os.path.join(output_dir, f"Toolify_AI_Revenue_{language.upper()}_{date_str}_analyzed_{timestamp}.xlsx")
                print(f"备用文件也被占用，使用时间戳文件名: {timestamp_output_file}")
                df.to_excel(timestamp_output_file, index=False)
                output_file = timestamp_output_file

        # 使用openpyxl美化Excel
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        sheet = wb.active
        format_excel_with_beauty(wb, sheet)
        wb.save(output_file)

        print(f"✅ 已将分析结果保存到 {output_file}")

        # 清理旧的Excel文件
        clean_old_excel_files(output_dir, language, date_str)

        return True
    except Exception as e:
        print(f"❌ 插入Markdown文件时出错: {str(e)}")
        traceback.print_exc()
        return False


def clean_old_excel_files(output_dir, language, date_str):
    """
    清理旧的Excel文件，只保留最新生成的标准文件
    """
    # 查找目录中的所有Excel文件
    pattern = f"Toolify_AI_Revenue_{language.upper()}_{date_str}_*.xlsx"
    excel_files = glob.glob(os.path.join(output_dir, pattern))

    # 最新生成的标准文件的名称
    standard_file = os.path.join(output_dir, f"Toolify_AI_Revenue_{language.upper()}_{date_str}_analyzed.xlsx")

    # 过滤要删除的文件
    files_to_delete = []
    for file in excel_files:
        # 如果是最新生成的标准文件，则跳过
        if file == standard_file:
            continue
        # 其他所有文件都添加到要删除的列表中
        files_to_delete.append(file)

    # 删除文件
    for file in files_to_delete:
        try:
            os.remove(file)
            print(f"✅ 已删除旧文件: {file}")
        except Exception as e:
            print(f"⚠️ 删除文件 {file} 时出错: {str(e)}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='数据爬取与分析工具')
    parser.add_argument('--scraping', dest='do_scraping',
                        action='store_true',
                        help='执行数据爬取步骤')
    parser.add_argument('--scrape', dest='scrape_lang',
                        choices=['cn', 'en', 'both'],
                        default='both',
                        help='爬取榜单的语言(cn=中文, en=英文, both=两者)')
    parser.add_argument('--analysis', dest='do_analysis',
                        action='store_true',
                        help='执行数据分析步骤')
    parser.add_argument('--analyze-products', dest='analyze_products',
                        action='store_true',
                        help='分析产品数据')
    parser.add_argument('--gpu', dest='use_gpu',
                        action='store_true',
                        help='使用GPU进行分析')
    parser.add_argument('--output-dir', dest='output_dir',
                        default='output',
                        help='分析结果输出目录')
    parser.add_argument('--batch-size', dest='batch_size',
                        type=int, default=10,
                        help='每批处理的数据量')
    parser.add_argument('-c', '--count', dest='count',
                        type=int, default=None,
                        help='要分析的产品数量')
    parser.add_argument('-l', '--language', dest='language',
                        choices=['cn', 'en', 'both'],
                        default='both',
                        help='爬取的语言(cn=中文, en=英文, both=两者)')
    parser.add_argument('--rank-range', dest='rank_range',
                        type=str, default=None,
                        help='分析指定排名范围的产品，如"1-5"')
    parser.add_argument('--product-ids', dest='product_ids',
                        type=str, default=None,
                        help='分析指定产品ID的产品，如"1,2,3"')
    parser.add_argument('--api', dest='api',
                        choices=['deepseek', 'openai'],
                        default='deepseek',
                        help='使用的API，默认为deepseek')
    parser.add_argument('--update-excel', dest='update_excel',
                        action='store_true',
                        help='将分析结果插入到Excel表格中')
    parser.add_argument('--retry', dest='retry_count',
                        type=int, default=3,
                        help='API调用失败时的重试次数，默认为3次')
    parser.add_argument('--insert-md', dest='insert_md',
                        action='store_true',
                        help='仅将已有的MD文件内容插入到Excel表格中，不进行新的分析')
    parser.add_argument('--excel-file', dest='excel_file',
                        help='Excel文件路径，如果不提供则自动查找')
    parser.add_argument('--markdown-dir', dest='markdown_dir',
                        help='Markdown文件目录，如果不提供则自动查找')
    parser.add_argument('--date', dest='date',
                        help='日期字符串，格式为YYYYMMDD，如果不提供则使用当前日期')
    parser.add_argument('--no-scraping', dest='no_scraping',
                        action='store_true',
                        help='不使用爬虫，直接使用本地Excel文件进行分析')
    parser.add_argument('--no-analysis', dest='do_analysis',
                        action='store_false',
                        help='跳过数据分析步骤')

    # 解析参数
    args = parser.parse_args()

    # 设置日期字符串
    date_str = args.date or datetime.now().strftime("%Y%m%d")

    # 检查GPU可用性
    if args.use_gpu and not GPU_AVAILABLE:
        print("⚠️ GPU不可用，将使用CPU")

    # 创建输出目录
    os.makedirs(args.output_dir, exist_ok=True)

    # 如果仅插入MD文件内容到Excel表格中
    if args.insert_md:
        print(f"\n🔍 将MD文件内容插入到Excel表格中")

        # 确定语言
        if args.language == 'both':
            languages = ["cn", "en"]
        else:
            languages = [args.language]

        # 确定日期字符串
        date_str = args.date or datetime.now().strftime("%Y%m%d")

        for lang in languages:
            print(f"\n🔍 开始处理{lang.upper()}语言的MD文件")

            # 插入MD文件内容到Excel表格中
            success = insert_md_to_excel(
                excel_file=args.excel_file,
                markdown_dir=args.markdown_dir,
                date_str=date_str,
                language=lang
            )

            if success:
                print("✅ 成功将MD文件内容插入到Excel表格中")
            else:
                print("❌ 插入MD文件内容失败")

        print("\n🎉 所有任务完成!")
        return

    # 初始化结果
    scraping_results = None
    analysis_result = None
    data_files = {}

    try:
        # 步骤1: 数据爬取
        if args.do_scraping:
            print("\n🔍 步骤1: 爬取Toolify数据")
            # 确保爬虫数据保存到toolify_data子目录
            scraping_dir = os.path.join(args.output_dir, "toolify_data")

            # 根据用户选择设置语言
            if args.language == 'both':
                languages = ["cn", "en"]
            else:
                languages = [args.language]

            # 执行爬取操作
            scraping_results = run_scraping(output_dir=scraping_dir, languages=languages)

            # 提取文件路径
            if scraping_results:
                for lang, result in scraping_results.items():
                    if "file" in result:
                        data_files[lang] = result["file"]
        else:
            print("\n🔍 已跳过数据爬取步骤")

            # 如果跳过爬取，尝试查找最新的文件
            data_dir = os.path.join(args.output_dir, "toolify_data")
            latest_files = find_latest_files(data_dir, "Toolify_AI_Revenue_*.xlsx")
            if latest_files:
                data_files = latest_files
                print(f"✅ 找到以下数据文件:")
                for lang, file_path in data_files.items():
                    print(f"  • {lang.upper()}: {file_path}")

        # 步骤2: 数据分析
        if args.do_analysis:
            print("\n🔍 步骤2: 分析Toolify数据")
            analysis_result = run_analysis(
                scraping_results=scraping_results,
                output_dir=args.output_dir,
                date_str=date_str
            )

            if analysis_result:
                print("✅ 数据分析完成")
            else:
                print("⚠️ 数据分析未完成或未返回结果")
        else:
            print("\n🔍 已跳过数据分析步骤")

        # 步骤3: 产品分析（可选）
        if args.analyze_products:
            print("\n🔍 步骤3: 分析产品详情")

            if not data_files:
                print("⚠️ 未找到数据文件，无法进行产品分析")
            else:
                success = run_product_analyzer(
                    data_files=data_files,
                    output_dir=f"output/toolify_analysis_{date_str}",
                    batch_size=args.batch_size,
                    use_gpu=args.use_gpu and GPU_AVAILABLE,
                    count=args.count
                )

                if success:
                    print("✅ 产品分析完成")
                else:
                    print("⚠️ 产品分析未完全成功")

        # 步骤4: 分析指定排名范围或产品ID的产品（可选）
        if args.rank_range or args.product_ids:
            print("\n🔍 步骤4: 分析指定产品")

            # 处理language参数
            if args.language == 'both':
                # 如果指定了排名范围但language是both，默认使用中文
                print("注意: 使用--rank-range时，将默认使用中文数据 (cn)")
                lang = "cn"
            else:
                lang = args.language

            # 设置输出目录
            output_dir_with_date = os.path.join(args.output_dir, f"toolify_analysis_{date_str}")
            os.makedirs(output_dir_with_date, exist_ok=True)

            # 如果指定了Excel文件路径，直接使用
            if args.excel_file and os.path.exists(args.excel_file):
                excel_file = args.excel_file
                print(f"\n🔍 使用指定的Excel文件: {excel_file}")
                try:
                    # 读取数据
                    tools = pd.read_excel(excel_file).to_dict('records')
                    print(f"✅ 成功读取 {len(tools)} 条数据")
                except Exception as e:
                    print(f"⚠️ 读取Excel文件时出错: {str(e)}")
                    traceback.print_exc()
                    tools = None
            # 如果没有指定文件路径但设置了no_scraping，尝试查找本地文件
            elif args.no_scraping:
                # 尝试查找本地文件
                excel_file = os.path.join("output", "toolify_data", f"Toolify_AI_Revenue_{lang.upper()}_{date_str}.xlsx")
                if not os.path.exists(excel_file):
                    # 尝试其他可能的文件名
                    alt_excel_file = os.path.join("output", "toolify_data", f"Toolify_Top_AI_Revenue_Rankings_{lang.upper()}_{date_str}.xlsx")
                    if os.path.exists(alt_excel_file):
                        excel_file = alt_excel_file
                    else:
                        # 尝试查找任何匹配的Excel文件
                        excel_files = glob.glob(os.path.join("output", "toolify_data", f"*{lang.upper()}*{date_str}*.xlsx"))
                        if excel_files:
                            excel_file = excel_files[0]
                        else:
                            print(f"⚠️ 找不到本地Excel文件，请指定 --excel-file 参数")
                            return

                print(f"\n🔍 使用本地Excel文件: {excel_file}")
                try:
                    # 读取数据
                    tools = pd.read_excel(excel_file).to_dict('records')
                    print(f"✅ 成功读取 {len(tools)} 条数据")
                except Exception as e:
                    print(f"⚠️ 读取Excel文件时出错: {str(e)}")
                    traceback.print_exc()
                    tools = None
            # 如果没有指定文件路径且没有设置no_scraping，使用爬虫
            else:
                # 先调用toolify_scraper.py中的main函数爬取最新的榜单数据
                print("\n🔍 正在调用toolify_scraper.py中的main函数爬取最新的榜单数据...")
                try:
                    scrape_main()
                    print("✅ 爬取完成")

                    # 设置输出文件名
                    excel_file = os.path.join("output", "toolify_data", f"Toolify_AI_Revenue_{lang.upper()}_{date_str}.xlsx")

                    # 检查文件是否存在
                    if os.path.exists(excel_file):
                        # 读取数据
                        tools = pd.read_excel(excel_file).to_dict('records')
                    else:
                        tools = None
                        print(f"⚠️ 找不到文件: {excel_file}")
                except Exception as e:
                    print(f"⚠️ 爬取数据时出错: {str(e)}")
                    traceback.print_exc()
                    tools = None
                    excel_file = None

            if not tools or not excel_file:
                print("⚠️ 获取榜单数据失败，无法进行分析")
            else:
                if args.no_scraping:
                    print(f"✅ 成功加载 {len(tools)} 个产品数据")
                else:
                    print(f"✅ 成功爬取 {len(tools)} 个产品数据")

                # 分析指定排名范围或产品ID的产品
                if args.rank_range:
                    print(f"\n🔍 开始分析指定排名范围的产品: {args.rank_range}")
                    success = analyze_specific_ranks(
                        args.rank_range,
                        lang,
                        date_str,
                        api=args.api,
                        use_gpu=args.use_gpu,
                        update_excel=args.update_excel,
                        retry_count=args.retry_count
                    )

                    if success:
                        print("✅ 指定排名范围的产品分析完成")
                    else:
                        print("⚠️ 指定排名范围的产品分析未完成或未返回结果")
                elif args.product_ids:
                    print(f"\n🔍 开始分析指定产品ID的产品: {args.product_ids}")
                    success = analyze_specific_products(
                        args.product_ids,
                        lang,
                        date_str,
                        api=args.api,
                        use_gpu=args.use_gpu,
                        update_excel=args.update_excel,
                        retry_count=args.retry_count,
                        excel_file=args.excel_file
                    )

                    if success:
                        print("✅ 指定产品ID的产品分析完成")
                    else:
                        print("⚠️ 指定产品ID的产品分析未完成或未返回结果")

        print("\n🎉 所有任务完成!")

    except Exception as e:
        print(f"\n❌ 程序执行出错: {str(e)}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()


