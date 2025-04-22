#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Toolify AI分析工具集 - 工具模块
提供共享的功能函数和类，减少代码重复，提高代码内聚性
"""

import os
import time
import pandas as pd
import numpy as np
from datetime import datetime
import re
import json
from typing import List, Dict, Any, Union, Tuple, Optional
import sys
import glob
import logging
import traceback
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("toolify_utils")

# 尝试加载环境变量
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("⚠️ python-dotenv未安装，无法加载.env文件")

# TensorFlow和GPU相关
TF_AVAILABLE = False
GPU_AVAILABLE = False

# 尝试导入TensorFlow并检查GPU
try:
    import tensorflow as tf

    # 设置TensorFlow日志级别
    os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # 0=INFO, 1=WARNING, 2=ERROR, 3=FATAL
    tf.get_logger().setLevel('ERROR')

    # 设置内存增长
    physical_devices = tf.config.list_physical_devices('GPU')
    if physical_devices:
        for device in physical_devices:
            tf.config.experimental.set_memory_growth(device, True)

    # 标记TensorFlow可用
    TF_AVAILABLE = True

    # 检查GPU可用性
    try:
        def get_available_gpus():
            """获取可用的GPU设备列表"""
            try:
                from tensorflow.python.client import device_lib
                local_device_protos = device_lib.list_local_devices()
                # 使用更通用的方式访问设备属性
                gpu_devices = []
                for device in local_device_protos:
                    if hasattr(device, 'device_type') and getattr(device, 'device_type') == 'GPU':
                        # 使用getattr安全地获取name属性
                        device_name = getattr(device, 'name', str(device))
                        gpu_devices.append(device_name)
                return gpu_devices
            except ImportError:
                logger.warning("⚠️ 无法导入device_lib模块，GPU检测将不可用")
                return []

        # 检查是否有可用的GPU
        gpus = get_available_gpus()
        GPU_AVAILABLE = len(gpus) > 0

        if GPU_AVAILABLE:
            logger.info(f"✅ TensorFlow已加载，发现{len(gpus)}个GPU设备")
            for i, gpu in enumerate(gpus):
                logger.info(f"  GPU #{i+1}: {gpu}")
        else:
            logger.info("✅ TensorFlow已加载，但未发现GPU设备")
    except Exception as e:
        logger.error(f"❌ 检查GPU可用性时出错: {str(e)}")
        GPU_AVAILABLE = False

except ImportError:
    logger.warning("⚠️ 未找到TensorFlow，某些功能将不可用")
    TF_AVAILABLE = False
    GPU_AVAILABLE = False
except Exception as e:
    logger.error(f"❌ 初始化TensorFlow时出错: {str(e)}")
    logger.debug(traceback.format_exc())
    TF_AVAILABLE = False
    GPU_AVAILABLE = False

# 加载分析框架
def load_analysis_framework():
    """从文件加载分析框架"""
    try:
        framework_file = "analysis_framework.txt"
        if os.path.exists(framework_file):
            with open(framework_file, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            print(f"⚠️ 未找到分析框架文件 {framework_file}，将使用默认框架")
            return """
## 产品分析框架

💡 这个产品解决的是什么问题？

👤 用户是谁？

🤔 用户为什么需要它？

🗣️ 用户是如何评价它的？

🔍 它是如何找到用户的？

💰 它赚钱吗？多少？

🧠 我从这个产品身上学到了什么？

🤔 它的什么做法不容易？

🤗 一句话推销：

💡 不同的方法：

🎉 我能做出来吗？

🧭 如何找到用户？

🤔 为什么是我？

❤️ 我能坚持吗？
"""
    except Exception as e:
        print(f"❌ 加载分析框架文件时出错: {str(e)}，将使用默认框架")
        return """## 产品分析框架\n\n💡 这个产品解决的是什么问题？\n\n👤 用户是谁？\n\n..."""

# 加载Markdown模板
def load_markdown_template():
    """从文件加载Markdown模板"""
    # 不再尝试加载外部文件，直接使用默认模板
    default_template = """# {name}

## 产品信息

- 📊 排名: {rank}
- 💰 收入: {revenue}
- 🔗 链接: [{website}]({url})
- 👀 月访问量: {monthly_visits}

## 产品描述

{description}

## 分析

{analysis}

## 评分卡

| 评估维度 | 得分(1-10) |
|---------|-----------|
| 创新性   | {innovation_score} |
| 商业模式可行性 | {business_model_score} |
| 市场增长潜力 | {growth_potential_score} |
| 总分 | {total_score} |

## SWOT分析

| 优势 (Strengths) | 劣势 (Weaknesses) |
|-----------------|------------------|
| {strengths} | {weaknesses} |

| 机会 (Opportunities) | 威胁 (Threats) |
|---------------------|--------------|
| {opportunities} | {threats} |
"""
    print(f"[INFO] 使用默认Markdown模板")
    return default_template

# 提取关键点
def extract_key_points(analysis):
    """从分析文本中提取关键点"""
    # 不再需要任何关键点，返回空字典
    return {}

# Excel处理函数
def format_excel_output(df, output_file):
    """改进Excel输出格式，使其更易于阅读"""
    try:
        # 首先尝试使用基本模式保存，确保至少有一个输出文件
        basic_output = output_file.replace('.xlsx', '_basic.xlsx')
        df.fillna('').to_excel(basic_output, index=False)
        print(f"✅ 已保存基本Excel文件: {basic_output}")

        # 尝试使用xlsxwriter引擎以支持格式设置
        try:
            # 使用xlsxwriter引擎以支持格式设置
            writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
            df.to_excel(writer, index=False, sheet_name='产品分析')

            # 获取xlsxwriter对象
            workbook = writer.book
            worksheet = writer.sheets['产品分析']

            # 设置列宽
            col_widths = {
                'A': 10,  # 排名
                'B': 25,  # 工具名称
                'C': 30,  # 工具链接
                'D': 25,  # 网站
                'E': 15,  # 支付平台
                'F': 15,  # 月访问量
                'G': 40,  # 描述
                'H': 15,  # 收入
                'I': 80   # 完整分析
            }

            # 安全地设置列宽
            for col, width in col_widths.items():
                try:
                    worksheet.set_column(f'{col}:{col}', width)
                except:
                    pass

            # 安全地关闭writer
            try:
                writer.close()
                print(f"✅ 已保存格式优化的Excel文件: {output_file}")
                return True
            except Exception as close_error:
                print(f"⚠️ 关闭Excel文件时出错: {str(close_error)}")
                return True  # 仍然返回成功，因为基本文件已保存

        except Exception as format_error:
            print(f"⚠️ 使用xlsxwriter格式化Excel时出错: {str(format_error)}")
            return True  # 仍然返回成功，因为基本文件已保存

    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {str(e)}")
        return False

# 清理URL函数
def clean_url(url):
    """
    清理URL，移除协议前缀和结尾的斜杠

    Args:
        url (str): 待清理的URL

    Returns:
        str: 清理后的URL
    """
    if url is None:
        return ""

    # 确保URL是字符串类型
    if not isinstance(url, str):
        try:
            url = str(url)
        except:
            return ""

    url = url.strip()

    # 移除HTTP/HTTPS协议
    if url.startswith("http://"):
        url = url[7:]
    elif url.startswith("https://"):
        url = url[8:]

    # 移除末尾的斜杠
    while url.endswith("/"):
        url = url[:-1]

    return url

# 从URL提取域名
def extract_domain(url):
    """
    从URL中提取域名

    Args:
        url (str): 输入URL

    Returns:
        str: 提取的域名
    """
    if not url:
        return ""

    # 确保URL是字符串类型
    if not isinstance(url, str):
        try:
            url = str(url)
        except:
            return ""

    # 清理URL
    url = clean_url(url)

    # 提取域名
    if not url:
        return ""

    parts = url.split("/")
    if not parts:
        return ""

    domain = parts[0]

    # 移除www.前缀
    if domain.startswith("www."):
        domain = domain[4:]

    return domain

# 查找最新的数据文件
def find_latest_files(directory: str, pattern: str = "Toolify_AI_Revenue_*.xlsx") -> Dict[str, str]:
    """
    查找目录中匹配给定模式的最新文件

    Args:
        directory (str): 要搜索的目录
        pattern (str): 文件名模式

    Returns:
        dict: 包含最新文件的字典，格式为 {"cn": "path/to/file.xlsx", "en": "path/to/file.xlsx"}
    """
    if not os.path.exists(directory) or not os.path.isdir(directory):
        logger.warning(f"目录不存在或不是目录: {directory}")
        return {}

    # 获取所有匹配的文件
    # 在多个可能的位置查找文件
    search_dirs = [
        directory,  # 原始目录
        os.path.join(directory, "toolify_data"),  # toolify_data子目录
        os.path.join("output", "toolify_data")  # 标准目录
    ]

    files = []
    for search_dir in search_dirs:
        if os.path.exists(search_dir) and os.path.isdir(search_dir):
            logger.info(f"在 {search_dir} 中查找文件...")
            found_files = glob.glob(os.path.join(search_dir, pattern))
            if found_files:
                files.extend(found_files)
                logger.info(f"在 {search_dir} 中找到 {len(found_files)} 个文件")

    if not files:
        logger.warning(f"在{directory}中未找到匹配{pattern}的文件")
        return {}

    # 按语言分类文件
    cn_files = [f for f in files if "_CN_" in f.upper()]
    en_files = [f for f in files if "_EN_" in f.upper()]

    # 按修改时间排序
    cn_files.sort(key=os.path.getmtime, reverse=True)
    en_files.sort(key=os.path.getmtime, reverse=True)

    result = {}

    # 获取最新的CN文件
    if cn_files:
        result["cn"] = cn_files[0]
        logger.info(f"找到最新的CN文件: {cn_files[0]}")

    # 获取最新的EN文件
    if en_files:
        result["en"] = en_files[0]
        logger.info(f"找到最新的EN文件: {en_files[0]}")

    return result

def markdown_to_plaintext(markdown_content):
    """
    将Markdown格式的文本转换为适合Excel单元格的纯文本格式

    Args:
        markdown_content (str): Markdown格式的文本内容

    Returns:
        str: 转换后的纯文本内容
    """
    if not markdown_content:
        return ""

    # 替换所有表格相关符号
    text = re.sub(r'\|', ' | ', markdown_content)

    # 替换标题格式（# 标题）
    text = re.sub(r'^\s*#{1,6}\s+(.+)$', r'\1', text, flags=re.MULTILINE)

    # 替换加粗和斜体格式
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)  # 加粗
    text = re.sub(r'__(.+?)__', r'\1', text)      # 加粗
    text = re.sub(r'\*(.+?)\*', r'\1', text)      # 斜体
    text = re.sub(r'_(.+?)_', r'\1', text)        # 斜体

    # 替换列表格式（- 和 * 开头的项）
    text = re.sub(r'^\s*[-*]\s+(.+)$', r'• \1', text, flags=re.MULTILINE)

    # 替换链接格式 [文本](链接)
    text = re.sub(r'\[(.+?)\]\((.+?)\)', r'\1 (\2)', text)

    # 替换图片格式 ![alt](链接)
    text = re.sub(r'!\[(.+?)\]\((.+?)\)', r'[图片: \1]', text)

    # 移除水平线
    text = re.sub(r'^\s*[-*_]{3,}\s*$', '', text, flags=re.MULTILINE)

    # 移除表格格式符号（| --- | 类的行）
    text = re.sub(r'^\s*\|[-:\s|]+\|\s*$', '', text, flags=re.MULTILINE)

    # 替换多个连续空行为单个空行
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text

# 更新Excel中的分析结果
def update_excel_with_analysis(excel_file, analysis_results, markdown_dir="output/toolify_analysis/markdown_files"):
    """更新Excel文件，移除产品分析列，只保留完整分析"""
    try:
        df = pd.read_excel(excel_file)

        # 确保移除产品分析列
        if "产品分析" in df.columns:
            df = df.drop(columns=["产品分析"])

        # 只保留完整分析列
        if "完整分析" not in df.columns:
            df["完整分析"] = ""

        # 更新分析结果
        total_results = len(analysis_results)
        print(f"\n开始更新Excel文件 [共{total_results}个产品]")

        # 定义进度条颜色和符号
        bar_length = 40
        filled_char = "█"  # 实心方块
        empty_char = "░"   # 空心方块

        for i, result in enumerate(analysis_results):
            # 计算进度
            progress = (i + 1) / total_results
            filled_length = int(bar_length * progress)

            # 清除当前行并显示新进度条
            if i > 0:  # 不是第一个元素
                print("\r", end="")

            # 打印进度条
            progress_bar = filled_char * filled_length + empty_char * (bar_length - filled_length)
            print(f"\r进度: {progress_bar} {progress:.1%} ({i+1}/{total_results})", end="")

            product = result["product"]
            tool_name = product.get("Tool Name", "") or product.get("工具名称", "")

            # 查找对应的行
            mask = None
            if "Tool Name" in df.columns and tool_name:
                mask = df["Tool Name"] == tool_name
            elif "工具名称" in df.columns and tool_name:
                mask = df["工具名称"] == tool_name

            if mask is None or not mask.any():
                continue

            # 获取分析工具信息
            api_name = result.get("api_name", "DeepSeek AI")  # 默认使用DeepSeek AI

            # 更新完整分析列
            if result["markdown_path"] and os.path.exists(result["markdown_path"]):
                try:
                    with open(result["markdown_path"], 'r', encoding='utf-8') as md_file:
                        markdown_content = md_file.read()

                        # 更新分析工具信息
                        if "🤖 分析工具: " in markdown_content:
                            markdown_content = re.sub(
                                r'🤖 分析工具: [^\n]+',
                                f'🤖 分析工具: {api_name}',
                                markdown_content
                            )
                        elif "🤖 Analysis Tool: " in markdown_content:
                            markdown_content = re.sub(
                                r'🤖 Analysis Tool: [^\n]+',
                                f'🤖 Analysis Tool: {api_name}',
                                markdown_content
                            )

                        # 转换为纯文本并更新Excel
                        plain_text = markdown_to_plaintext(markdown_content)
                        df.loc[mask, "完整分析"] = plain_text

                        # 将更新后的内容写回文件
                        with open(result["markdown_path"], 'w', encoding='utf-8') as md_out:
                            md_out.write(markdown_content)
                except Exception as md_error:
                    print(f"\n⚠️ 读取markdown文件失败: {str(md_error)}")

        # 完成后打印完成信息
        print(f"\r进度: {filled_char * bar_length} 100.0% ({total_results}/{total_results}) ✅ 完成!\n")

        # 确定输出目录
        # 使用markdown_dir的父目录作为输出目录
        if markdown_dir:
            # 如果是类似 "output\toolify_analysis_20250419\cn\markdown_files" 的路径
            # 则输出目录应为 "output\toolify_analysis_20250419"
            parts = os.path.normpath(markdown_dir).split(os.sep)
            if len(parts) >= 3 and parts[0] == "output" and parts[1].startswith("toolify_analysis"):
                # 使用前两部分作为输出目录
                output_dir = os.path.join(parts[0], parts[1])
            else:
                # 如果不符合预期格式，则使用父目录
                output_dir = os.path.dirname(markdown_dir)
        else:
            # 默认输出目录
            output_dir = "output/toolify_analysis"

        # 确保目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 生成输出文件名
        file_name = os.path.basename(excel_file)
        base_name, ext = os.path.splitext(file_name)
        output_file = os.path.join(output_dir, f"{base_name}_analyzed{ext}")

        # 保存Excel文件
        format_excel_with_beauty(df, output_file)

        print(f"✅ 已将分析结果保存到Excel文件: {output_file}")
        return output_file

    except Exception as e:
        print(f"❌ 更新Excel文件失败: {str(e)}")
        return None

def format_excel_with_beauty(df, output_file):
    """用美观的格式保存Excel文件"""
    try:
        # 创建Excel写入器
        writer = pd.ExcelWriter(output_file, engine='openpyxl')

        # 写入数据
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 应用格式（使用openpyxl的样式）
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 设置列宽
        for idx, col in enumerate(df.columns):
            column_letter = get_column_letter(idx + 1)
            # 标准列宽
            col_width = 15

            # 特定列宽度调整
            if col in ["产品分析", "完整分析"]:
                col_width = 50  # 分析列宽一些
            elif col in ["收入", "Revenue", "年收入", "Annual Revenue"]:
                col_width = 12

            worksheet.column_dimensions[column_letter].width = col_width

        # 设置标题行格式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 应用标题行格式
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置单元格边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用到所有单元格
        for row in range(1, len(df) + 2):  # +2 因为有标题行和索引从1开始
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col_num)
                cell.border = thin_border

                # 内容单元格格式
                if row > 1:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 特殊列格式
                    column_name = df.columns[col_num-1]
                    if column_name in ["产品分析", "完整分析"]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 保存文件
        writer.close()
        print(f"✅ Excel文件已保存: {output_file}")
        return True
    except Exception as e:
        print(f"❌ 保存Excel格式化文件失败: {str(e)}")
        return False

# 常量
ANALYSIS_FRAMEWORK = load_analysis_framework()
MARKDOWN_TEMPLATE = load_markdown_template()

def get_datetime_str(format="%Y%m%d"):
    """
    获取格式化的日期时间字符串

    Args:
        format (str): 日期时间格式

    Returns:
        str: 格式化的日期时间字符串
    """
    return datetime.now().strftime(format)

def format_number(number):
    """
    格式化数字，添加千位分隔符

    Args:
        number: 要格式化的数字

    Returns:
        str: 格式化后的数字字符串
    """
    if number is None:
        return "N/A"

    try:
        return f"{int(number):,}"
    except (ValueError, TypeError):
        try:
            return f"{float(number):,.2f}"
        except (ValueError, TypeError):
            return str(number)

def convert_size(size_bytes):
    """
    将字节转换为人类可读的大小

    Args:
        size_bytes (int): 字节大小

    Returns:
        str: 人类可读的大小字符串
    """
    if size_bytes == 0:
        return "0B"

    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = 0

    while size_bytes >= 1024 and i < len(size_name) - 1:
        size_bytes /= 1024
        i += 1

    return f"{size_bytes:.2f} {size_name[i]}"

def parse_monthly_visits(visits_str):
    """
    解析月访问量字符串为数值

    Args:
        visits_str (str): 访问量字符串，例如 "1.2M" 或 "500K"

    Returns:
        float: 数值形式的访问量
    """
    if not visits_str or visits_str == "N/A":
        return 0.0

    # 移除所有空格和逗号
    visits_str = visits_str.strip().replace(" ", "").replace(",", "")

    # 提取数字和单位
    match = re.match(r"^([\d.]+)([KMB])?$", visits_str.upper())

    if not match:
        try:
            return float(visits_str)
        except (ValueError, TypeError):
            return 0.0

    value, unit = match.groups()
    factor = 1

    # 根据单位调整数值
    if unit == "K":
        factor = 1000
    elif unit == "M":
        factor = 1000000
    elif unit == "B":
        factor = 1000000000

    try:
        return float(value) * factor
    except (ValueError, TypeError):
        return 0.0

# 可在此处添加更多工具函数

def clean_number(value: str) -> Optional[float]:
    """
    清理并转换字符串数字为浮点数

    Args:
        value: 要转换的字符串值，可能包含逗号、货币符号等

    Returns:
        清理后的浮点数值，如果转换失败则返回None
    """
    if not value or not isinstance(value, str):
        return None

    # 移除常见非数字字符
    for char in ['$', '¥', '€', ',', ' ', '+']:
        value = value.replace(char, '')

    # 处理k/m/b简写
    multipliers = {
        'k': 1000,
        'm': 1000000,
        'b': 1000000000
    }

    value = value.lower().strip()
    for suffix, multiplier in multipliers.items():
        if value.endswith(suffix):
            try:
                return float(value[:-1]) * multiplier
            except ValueError:
                return None

    # 尝试直接转换为浮点数
    try:
        return float(value)
    except ValueError:
        return None

def format_filename(base_name: str, date_str: Optional[str] = None, extension: str = 'xlsx') -> str:
    """
    格式化文件名，添加日期后缀

    Args:
        base_name: 基本文件名
        date_str: 日期字符串，如果不提供则使用当前日期
        extension: 文件扩展名，默认为xlsx

    Returns:
        格式化后的文件名
    """
    if not date_str:
        date_str = datetime.now().strftime("%Y%m%d")

    # 确保扩展名格式正确
    ext = extension.lstrip('.')

    return f"{base_name}_{date_str}.{ext}"

def ensure_dir(directory: str) -> None:
    """
    确保目录存在，如果不存在则创建

    Args:
        directory: 要确保存在的目录路径
    """
    if not os.path.exists(directory):
        os.makedirs(directory)
        logger.info(f"创建目录: {directory}")

def get_display_path(file_path: str, max_length: int = 60) -> str:
    """
    获取用于显示的文件路径，如果太长则截断

    Args:
        file_path: 完整文件路径
        max_length: 最大显示长度

    Returns:
        适合显示的文件路径
    """
    if not file_path:
        return "[无路径]"

    if len(file_path) <= max_length:
        return file_path

    # 保留文件名和部分路径
    file_name = os.path.basename(file_path)
    dir_part = os.path.dirname(file_path)

    available_length = max_length - len(file_name) - 4  # 4 for ".../"

    if available_length <= 0:
        # 文件名已经太长，只显示部分文件名
        return f"...{file_name[-(max_length-3):]}"

    # 显示部分目录和完整文件名
    return f"...{dir_part[-available_length:]}/{file_name}"
